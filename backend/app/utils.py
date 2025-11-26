import os
import logging
from datetime import datetime, timezone

# Directory to store exported session JSONs (if needed)
# Use absolute path to avoid issues in Docker
SESSION_EXPORT_DIR = os.path.abspath(os.environ.get("SESSION_EXPORT_DIR", "session_exports"))
os.makedirs(SESSION_EXPORT_DIR, exist_ok=True)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Utility Functions ---

def safe_filename(name: str) -> str:
    """
    Sanitize a string to be filesystem-safe.
    Replaces all non-alphanumeric characters with underscores.
    """
    return "".join(c if c.isalnum() or c in ("_", "-") else "_" for c in name)

def timestamp_now(fmt: str = "%Y%m%d_%H%M%S") -> str:
    """
    Return a UTC timestamp string for filenames or logs.
    Default format: YYYYMMDD_HHMMSS
    """
    return datetime.utcnow().strftime(fmt)

def utc_now() -> datetime:
    """
    Return current UTC datetime with timezone info.
    """
    return datetime.now(timezone.utc)

def export_session_data(session_name: str, data: dict) -> str:
    """
    Export session data to a JSON file in SESSION_EXPORT_DIR.
    Returns the path of the saved file.
    """
    import json
    filename = f"{safe_filename(session_name)}_{timestamp_now()}.json"
    path = os.path.join(SESSION_EXPORT_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    logger.info(f"Session data exported to {path}")
    return path

def load_session_data(file_path: str) -> dict:
    """
    Load session data from a JSON file.
    """
    import json
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Session file '{file_path}' does not exist.")
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)

def export_session_csv(session_name: str, infringements: list, session_info: dict = None) -> str:
    """
    Export session data to CSV format.
    Returns the path of the saved file.
    """
    import csv
    filename = f"{safe_filename(session_name)}_{timestamp_now()}.csv"
    path = os.path.join(SESSION_EXPORT_DIR, filename)
    
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        
        # Write session info header
        if session_info:
            writer.writerow(["Session Information"])
            writer.writerow(["Name", session_info.get("name", "")])
            writer.writerow(["Status", session_info.get("status", "")])
            writer.writerow(["Started At", session_info.get("started_at", "")])
            writer.writerow([])  # Empty row
        
        # Write infringements header
        writer.writerow(["Infringements"])
        writer.writerow([
            "ID", "Kart Number", "Turn Number", "Description", "Observer",
            "Warning Count", "Penalty Due", "Penalty Description", "Penalty Taken", "Timestamp"
        ])
        
        # Write infringement data
        for inf in infringements:
            writer.writerow([
                inf.get("id", ""),
                inf.get("kart_number", ""),
                inf.get("turn_number", ""),
                inf.get("description", ""),
                inf.get("observer", ""),
                inf.get("warning_count", ""),
                inf.get("penalty_due", ""),
                inf.get("penalty_description", ""),
                inf.get("penalty_taken", ""),
                inf.get("timestamp", "")
            ])
        
        # Write history if available
        has_history = any(inf.get("history") for inf in infringements)
        if has_history:
            writer.writerow([])  # Empty row
            writer.writerow(["Infringement History"])
            writer.writerow([
                "Infringement ID", "Action", "Performed By", "Observer", "Details", "Timestamp"
            ])
            
            for inf in infringements:
                for hist in inf.get("history", []):
                    writer.writerow([
                        inf.get("id", ""),
                        hist.get("action", ""),
                        hist.get("performed_by", ""),
                        hist.get("observer", ""),
                        hist.get("details", ""),
                        hist.get("timestamp", "")
                    ])
    
    logger.info(f"Session data exported to CSV: {path}")
    return path

def export_session_excel(session_name: str, infringements: list, session_info: dict = None) -> str:
    """
    Export session data to Excel format (.xlsx).
    Returns the path of the saved file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    filename = f"{safe_filename(session_name)}_{timestamp_now()}.xlsx"
    path = os.path.join(SESSION_EXPORT_DIR, filename)
    
    wb = Workbook()
    
    # === Sheet 1: Infringements ===
    ws = wb.active
    ws.title = "Infringements"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write session info
    if session_info:
        ws["A1"] = "Session Information"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Name:"
        ws["B2"] = session_info.get("name", "")
        ws["A3"] = "Status:"
        ws["B3"] = session_info.get("status", "")
        ws["A4"] = "Started At:"
        ws["B4"] = session_info.get("started_at", "")
        ws.append([])  # Empty row
    
    # Helper function to format timestamp as hh:mm:ss
    def format_time(timestamp_str):
        if not timestamp_str:
            return ""
        try:
            # Parse ISO format timestamp
            dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            return dt.strftime("%H:%M:%S")
        except (ValueError, AttributeError):
            return str(timestamp_str) if timestamp_str else ""
    
    # Write infringements header
    start_row = 6 if session_info else 1
    headers = [
        "ID", "Kart Number", "Turn Number", "Description", "Observer",
        "Warning Count", "Penalty Due", "Penalty Description", "Penalty Taken", "Timestamp"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write infringement data
    for inf in infringements:
        timestamp_str = inf.get("timestamp", "")
        penalty_taken_str = inf.get("penalty_taken", "")
        
        row = [
            inf.get("id", ""),
            inf.get("kart_number", ""),
            inf.get("turn_number", ""),
            inf.get("description", ""),
            inf.get("observer", ""),
            inf.get("warning_count", ""),
            inf.get("penalty_due", ""),
            inf.get("penalty_description", ""),
            format_time(penalty_taken_str),
            format_time(timestamp_str)
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # === Sheet 2: History ===
    has_history = any(inf.get("history") for inf in infringements)
    if has_history:
        ws2 = wb.create_sheet("History")
        history_headers = [
            "Infringement ID", "Action", "Performed By", "Observer", "Details", "Timestamp"
        ]
        
        for col, header in enumerate(history_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for inf in infringements:
            for hist in inf.get("history", []):
                row = [
                    inf.get("id", ""),
                    hist.get("action", ""),
                    hist.get("performed_by", ""),
                    hist.get("observer", ""),
                    hist.get("details", ""),
                    format_time(hist.get("timestamp", ""))
                ]
                ws2.append(row)
        
        # Auto-adjust column widths for history sheet
        for col in range(1, len(history_headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws2[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(path)
    logger.info(f"Session data exported to Excel: {path}")
    return path

def import_session_excel(file_path: str) -> dict:
    """
    Import session data from an Excel file (.xlsx).
    Returns a dictionary with session_info, infringements, and history.
    
    Expected format:
    - Sheet "Infringements": Session info at top, then headers, then data rows
    - Sheet "History" (optional): Headers, then history data rows
    """
    from openpyxl import load_workbook
    from dateutil import parser as date_parser
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file '{file_path}' does not exist.")
    
    wb = load_workbook(file_path, data_only=True)
    
    # === Parse Sheet 1: Infringements ===
    if "Infringements" not in wb.sheetnames:
        raise ValueError("Excel file must contain a sheet named 'Infringements'")
    
    ws = wb["Infringements"]
    
    # Parse session info (rows 1-4)
    session_info = {}
    session_name = None
    
    # Look for session info in first few rows
    for row_idx in range(1, 6):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value
        
        if cell_a and isinstance(cell_a, str):
            if "name" in cell_a.lower() and cell_b:
                session_name = str(cell_b).strip()
                session_info["name"] = session_name
            elif "status" in cell_a.lower() and cell_b:
                session_info["status"] = str(cell_b).strip()
            elif "started" in cell_a.lower() and cell_b:
                try:
                    # Try to parse the date
                    if isinstance(cell_b, datetime):
                        session_info["started_at"] = cell_b.isoformat()
                    else:
                        session_info["started_at"] = date_parser.parse(str(cell_b)).isoformat()
                except:
                    session_info["started_at"] = str(cell_b)
    
    # Find the header row (look for "ID" or "Kart Number")
    header_row = None
    for row_idx in range(1, 20):  # Check first 20 rows
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value and str(cell_value).strip().upper() in ["ID", "KART NUMBER"]:
            header_row = row_idx
            break
    
    if not header_row:
        raise ValueError("Could not find header row in 'Infringements' sheet")
    
    # Read headers
    headers = []
    for col in range(1, 20):  # Check up to 20 columns
        cell_value = ws.cell(row=header_row, column=col).value
        if not cell_value:
            break
        headers.append(str(cell_value).strip())
    
    # Map headers to field names
    header_map = {
        "ID": "id",
        "Kart Number": "kart_number",
        "Turn Number": "turn_number",
        "Description": "description",
        "Observer": "observer",
        "Warning Count": "warning_count",
        "Penalty Due": "penalty_due",
        "Penalty Description": "penalty_description",
        "Penalty Taken": "penalty_taken",
        "Timestamp": "timestamp"
    }
    
    # Read infringement data
    infringements = []
    data_start_row = header_row + 1
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        # Check if row is empty
        first_cell = ws.cell(row=row_idx, column=1).value
        if not first_cell:
            continue
        
        inf = {}
        for col_idx, header in enumerate(headers, 1):
            field_name = header_map.get(header)
            if not field_name:
                continue
            
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            
            # Convert based on field type
            if field_name == "id":
                inf[field_name] = int(cell_value) if cell_value is not None else None
            elif field_name in ["kart_number", "warning_count"]:
                try:
                    inf[field_name] = int(cell_value) if cell_value is not None else None
                except (ValueError, TypeError):
                    inf[field_name] = None
            elif field_name == "turn_number":
                try:
                    inf[field_name] = str(cell_value).strip() if cell_value is not None else None
                except (ValueError, TypeError):
                    inf[field_name] = None
            elif field_name in ["penalty_taken", "timestamp"]:
                if cell_value:
                    try:
                        if isinstance(cell_value, datetime):
                            inf[field_name] = cell_value.isoformat()
                        else:
                            inf[field_name] = date_parser.parse(str(cell_value)).isoformat()
                    except:
                        inf[field_name] = str(cell_value) if cell_value else None
                else:
                    inf[field_name] = None
            elif field_name == "penalty_due":
                inf[field_name] = str(cell_value).strip() if cell_value else "No"
            else:
                inf[field_name] = str(cell_value).strip() if cell_value else None
        
        # Only add if we have at least kart_number and description
        if inf.get("kart_number") and inf.get("description"):
            infringements.append(inf)
    
    # === Parse Sheet 2: History (if exists) ===
    history = []
    if "History" in wb.sheetnames:
        ws2 = wb["History"]
        
        # Find header row
        history_header_row = None
        for row_idx in range(1, 10):
            cell_value = ws2.cell(row=row_idx, column=1).value
            if cell_value and "Infringement ID" in str(cell_value):
                history_header_row = row_idx
                break
        
        if history_header_row:
            # Read history headers
            history_headers = []
            for col in range(1, 20):
                cell_value = ws2.cell(row=history_header_row, column=col).value
                if not cell_value:
                    break
                history_headers.append(str(cell_value).strip())
            
            history_header_map = {
                "Infringement ID": "infringement_id",
                "Action": "action",
                "Performed By": "performed_by",
                "Observer": "observer",
                "Details": "details",
                "Timestamp": "timestamp"
            }
            
            # Read history data
            for row_idx in range(history_header_row + 1, ws2.max_row + 1):
                first_cell = ws2.cell(row=row_idx, column=1).value
                if not first_cell:
                    continue
                
                hist = {}
                for col_idx, header in enumerate(history_headers, 1):
                    field_name = history_header_map.get(header)
                    if not field_name:
                        continue
                    
                    cell_value = ws2.cell(row=row_idx, column=col_idx).value
                    
                    if field_name == "infringement_id":
                        try:
                            hist[field_name] = int(cell_value) if cell_value is not None else None
                        except (ValueError, TypeError):
                            hist[field_name] = None
                    elif field_name == "timestamp":
                        if cell_value:
                            try:
                                if isinstance(cell_value, datetime):
                                    hist[field_name] = cell_value.isoformat()
                                else:
                                    hist[field_name] = date_parser.parse(str(cell_value)).isoformat()
                            except:
                                hist[field_name] = str(cell_value) if cell_value else None
                        else:
                            hist[field_name] = None
                    else:
                        hist[field_name] = str(cell_value).strip() if cell_value else None
                
                if hist.get("infringement_id") and hist.get("action"):
                    history.append(hist)
    
    # Group history by infringement_id
    history_by_inf_id = {}
    for hist in history:
        inf_id = hist.get("infringement_id")
        if inf_id:
            if inf_id not in history_by_inf_id:
                history_by_inf_id[inf_id] = []
            history_by_inf_id[inf_id].append(hist)
    
    # Attach history to infringements
    for inf in infringements:
        inf_id = inf.get("id")
        if inf_id and inf_id in history_by_inf_id:
            inf["history"] = history_by_inf_id[inf_id]
        else:
            inf["history"] = []
    
    logger.info(f"Imported {len(infringements)} infringements and {len(history)} history records from {file_path}")
    
    return {
        "session_info": session_info,
        "infringements": infringements
    }

def import_session_csv(file_path: str) -> dict:
    """
    Import session data from a CSV file.
    Returns a dictionary with session_info, infringements, and history.
    
    Expected format matches export_session_csv output.
    """
    import csv
    from dateutil import parser as date_parser
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"CSV file '{file_path}' does not exist.")
    
    session_info = {}
    infringements = []
    history = []
    
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    # Parse session info (first few rows)
    i = 0
    while i < len(rows) and i < 10:
        row = rows[i]
        if len(row) >= 2:
            if row[0] and "name" in row[0].lower():
                session_info["name"] = row[1] if len(row) > 1 else ""
            elif row[0] and "status" in row[0].lower():
                session_info["status"] = row[1] if len(row) > 1 else ""
            elif row[0] and "started" in row[0].lower():
                session_info["started_at"] = row[1] if len(row) > 1 else ""
        if row and row[0] == "Infringements":
            i += 1
            break
        i += 1
    
    # Find header row
    header_row_idx = None
    for idx in range(i, min(i + 5, len(rows))):
        if rows[idx] and len(rows[idx]) > 0:
            first_col = rows[idx][0].strip() if rows[idx][0] else ""
            if first_col.upper() in ["ID", "KART NUMBER"]:
                header_row_idx = idx
                break
    
    if not header_row_idx:
        raise ValueError("Could not find infringements header row in CSV")
    
    # Read headers
    headers = [h.strip() for h in rows[header_row_idx]]
    
    # Map headers to field names
    header_map = {
        "ID": "id",
        "Kart Number": "kart_number",
        "Turn Number": "turn_number",
        "Description": "description",
        "Observer": "observer",
        "Warning Count": "warning_count",
        "Penalty Due": "penalty_due",
        "Penalty Description": "penalty_description",
        "Penalty Taken": "penalty_taken",
        "Timestamp": "timestamp"
    }
    
    # Read infringement data
    data_start = header_row_idx + 1
    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        
        # Skip empty rows
        if not row or not row[0]:
            # Check if we've hit the history section
            if row_idx < len(rows) - 1 and rows[row_idx + 1] and len(rows[row_idx + 1]) > 0:
                if rows[row_idx + 1][0] and "History" in rows[row_idx + 1][0]:
                    break
            continue
        
        # Check if we've hit the history section
        if row[0] and "History" in row[0]:
            break
        
        inf = {}
        for col_idx, header in enumerate(headers):
            if col_idx >= len(row):
                break
            field_name = header_map.get(header)
            if not field_name:
                continue
            
            value = row[col_idx].strip() if col_idx < len(row) and row[col_idx] else ""
            
            # Convert based on field type
            if field_name == "id":
                try:
                    inf[field_name] = int(value) if value else None
                except:
                    inf[field_name] = None
            elif field_name in ["kart_number", "warning_count"]:
                try:
                    inf[field_name] = int(value) if value else None
                except:
                    inf[field_name] = None
            elif field_name == "turn_number":
                inf[field_name] = value if value else None
            elif field_name in ["penalty_taken", "timestamp"]:
                if value:
                    try:
                        inf[field_name] = date_parser.parse(value).isoformat()
                    except:
                        inf[field_name] = value
                else:
                    inf[field_name] = None
            elif field_name == "penalty_due":
                inf[field_name] = value if value else "No"
            else:
                inf[field_name] = value if value else None
        
        # Only add if we have at least kart_number and description
        if inf.get("kart_number") and inf.get("description"):
            infringements.append(inf)
    
    # Parse history if present
    history_start = None
    for idx in range(len(rows)):
        if rows[idx] and len(rows[idx]) > 0 and "History" in rows[idx][0]:
            history_start = idx + 1
            break
    
    if history_start:
        # Find history header
        history_header_idx = None
        for idx in range(history_start, min(history_start + 3, len(rows))):
            if rows[idx] and len(rows[idx]) > 0:
                if "Infringement ID" in rows[idx][0]:
                    history_header_idx = idx
                    break
        
        if history_header_idx:
            history_headers = [h.strip() for h in rows[history_header_idx]]
            history_header_map = {
                "Infringement ID": "infringement_id",
                "Action": "action",
                "Performed By": "performed_by",
                "Observer": "observer",
                "Details": "details",
                "Timestamp": "timestamp"
            }
            
            for row_idx in range(history_header_idx + 1, len(rows)):
                row = rows[row_idx]
                if not row or not row[0]:
                    continue
                
                hist = {}
                for col_idx, header in enumerate(history_headers):
                    if col_idx >= len(row):
                        break
                    field_name = history_header_map.get(header)
                    if not field_name:
                        continue
                    
                    value = row[col_idx].strip() if col_idx < len(row) and row[col_idx] else ""
                    
                    if field_name == "infringement_id":
                        try:
                            hist[field_name] = int(value) if value else None
                        except:
                            hist[field_name] = None
                    elif field_name == "timestamp":
                        if value:
                            try:
                                hist[field_name] = date_parser.parse(value).isoformat()
                            except:
                                hist[field_name] = value
                        else:
                            hist[field_name] = None
                    else:
                        hist[field_name] = value if value else None
                
                if hist.get("infringement_id") and hist.get("action"):
                    history.append(hist)
    
    # Group history by infringement_id
    history_by_inf_id = {}
    for hist in history:
        inf_id = hist.get("infringement_id")
        if inf_id:
            if inf_id not in history_by_inf_id:
                history_by_inf_id[inf_id] = []
            history_by_inf_id[inf_id].append(hist)
    
    # Attach history to infringements
    for inf in infringements:
        inf_id = inf.get("id")
        if inf_id and inf_id in history_by_inf_id:
            inf["history"] = history_by_inf_id[inf_id]
        else:
            inf["history"] = []
    
    logger.info(f"Imported {len(infringements)} infringements and {len(history)} history records from CSV {file_path}")
    
    return {
        "session_info": session_info,
        "infringements": infringements
    }
